#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
SuccessFactors API Documentation Generator

This script generates SuccessFactors EC metadata, queries API endpoints,
parses responses, looks up field attributes, and generates an Excel documentation file.
"""

import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import json
import xml.etree.ElementTree as ET
import pandas as pd
import sys
import traceback
from datetime import datetime

# Import configuration from config_SF.py
try:
    from config_SF import *
except ImportError:
    print("Error: config_SF.py file not found or contains errors.")
    print("Please ensure config_SF.py exists in the same directory as this script.")
    sys.exit(1)

# =============================================================================
# OAuth 2.0 Configuration (Simulation Environment)
# =============================================================================

SF_CLIENT_ID = os.getenv('SF_CLIENT_ID', 'YOUR_CLIENT_ID')
SF_USER_ID = os.getenv('SF_USER_ID', 'YOUR_USER_ID')
SF_COMPANY_ID = os.getenv('SF_COMPANY_ID', 'YOUR_COMPANY_ID')
SF_TOKEN_URL = os.getenv('SF_TOKEN_URL', 'https://apidemo.sapsf.com/oauth/token')
SF_IDP_URL = os.getenv('SF_IDP_URL', 'https://apidemo.sapsf.com/oauth/idp')
SF_PRIVATE_KEY = os.getenv("SF_PRIVATE_KEY", "")  # Load from environment or secret management service

# Global variable to store the dynamic token
DYNAMIC_ACCESS_TOKEN = None

# =============================================================================
# Authentication Helper Function
# =============================================================================

def get_assertion():
    """Get SAML assertion from SuccessFactors."""
    print("Generating SAML assertion...")
    payload = {
        "client_id": SF_CLIENT_ID,
        "user_id": SF_USER_ID,
        "token_url": SF_TOKEN_URL,
        "private_key": SF_PRIVATE_KEY
    }
    try:
        response = requests.post(SF_IDP_URL, data=payload)
        response.raise_for_status()
        return response.text
    except Exception as e:
        print(f"Error getting assertion: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response: {e.response.text}")
        return None

def get_access_token(assertion):
    """Get OAuth 2.0 Access Token using assertion."""
    print("Getting Access Token...")
    payload = {
        "company_id": SF_COMPANY_ID,
        "client_id": SF_CLIENT_ID,
        "grant_type": "urn:ietf:params:oauth:grant-type:saml2-bearer",
        "user_id": SF_USER_ID,
        "assertion": assertion,
        "new_token": "true"
    }
    try:
        response = requests.post(SF_TOKEN_URL, data=payload)
        response.raise_for_status()
        data = response.json()
        return data.get("access_token")
    except Exception as e:
        print(f"Error getting access token: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response: {e.response.text}")
        return None

def make_request(url, method='GET', **kwargs):
    """Make HTTP request with Dynamic Bearer token, fallback to Config Bearer, then Basic auth."""
    
    # 1. Try Dynamic Access Token
    if DYNAMIC_ACCESS_TOKEN:
        headers = kwargs.get('headers', {})
        headers['Authorization'] = f'Bearer {DYNAMIC_ACCESS_TOKEN}'
        kwargs['headers'] = headers
        if 'verify' not in kwargs:
            kwargs['verify'] = True
        
        try:
            if method.upper() == 'GET':
                response = requests.get(url, **kwargs)
            elif method.upper() == 'POST':
                response = requests.post(url, **kwargs)
            else:
                response = requests.request(method, url, **kwargs)
            
            if response.status_code not in [401, 403]:
                return response
            print(f"Dynamic Bearer token failed with {response.status_code}.")
        except Exception as e:
            print(f"Dynamic Bearer token request failed: {e}.")

    # 2. Try Config Bearer Token
    if BEARER_TOKEN:
        headers = kwargs.get('headers', {})
        headers['Authorization'] = f'Bearer {BEARER_TOKEN}'
        kwargs['headers'] = headers
        if 'verify' not in kwargs:
            kwargs['verify'] = True
        
        try:
            if method.upper() == 'GET':
                response = requests.get(url, **kwargs)
            elif method.upper() == 'POST':
                response = requests.post(url, **kwargs)
            else:
                response = requests.request(method, url, **kwargs)
            
            if response.status_code not in [401, 403]:
                return response
            print(f"Config Bearer token failed with {response.status_code}.")
        except Exception as e:
            print(f"Config Bearer token request failed: {e}.")

    # 3. Fallback to Basic Auth
    if 'headers' in kwargs and 'Authorization' in kwargs['headers']:
        del kwargs['headers']['Authorization']
    
    kwargs['auth'] = HTTPBasicAuth(USERNAME, PASSWORD)
    if 'verify' not in kwargs:
        kwargs['verify'] = True
        
    if method.upper() == 'GET':
        return requests.get(url, **kwargs)
    elif method.upper() == 'POST':
        return requests.post(url, **kwargs)
    else:
        return requests.request(method, url, **kwargs)

# =============================================================================
# XML Metadata Parsing Helpers
# =============================================================================

def get_text(node, tag):
    el = node.find(tag, ns)
    return el.text if el is not None else "Null"

def get_sap_tags(doc):
    tags = doc.find('.//sap:tagcollection', ns)
    if tags is not None:
        return ', '.join([t.text for t in tags.findall('sap:tag', ns)])
    return "Null"

def get_attr_value(node, attr):
    if attr in node.attrib:
        return node.attrib[attr]
    if 'sap:' + attr in node.attrib:
        return node.attrib['sap:' + attr]
    sap_ns = '{http://www.successfactors.com/edm/sap}' + attr
    if sap_ns in node.attrib:
        return node.attrib[sap_ns]
    return "Null"

def clean_attr(attr):
    return re.sub(r'^\{.*\}', '', attr).replace('sap:', '')


# =============================================================================
# EC Metadata Extraction
# =============================================================================

def extract_ec_odata_api_dictionary():
    """Extract API metadata for each entity and create data dictionaries."""
    print("Extracting EC OData API Dictionary...")
    
    # Download and parse metadata for each entity set
    metadata_trees = {}
    for entity in ENTITY_SETS:
        url = f"https://{API_SERVER}/odata/v2/{entity}/$metadata"
        print(f"Fetching metadata for {entity}...")
        try:
            response = make_request(url)
            if response.status_code == 200:
                metadata_trees[entity] = ET.ElementTree(ET.fromstring(response.content))
            else:
                print(f"Failed to fetch metadata for {entity}: {response.status_code}")
        except Exception as e:
            print(f"Error fetching metadata for {entity}: {e}")

    # Collect detailed field metadata
    rows = []
    all_attrs = set()

    for tree in metadata_trees.values():
        root = tree.getroot()
        for schema in root.findall('.//{http://schemas.microsoft.com/ado/2008/09/edm}Schema'):
            if schema.attrib.get('Namespace') == 'SFOData':
                for et in schema.findall('EntityType', ns):
                    for prop in et.findall('Property', ns):
                        all_attrs.update([clean_attr(a) for a in prop.attrib.keys()])
                    for nav in et.findall('NavigationProperty', ns):
                        all_attrs.update([clean_attr(a) for a in nav.attrib.keys()])

    all_attrs = sorted(all_attrs)
    all_attrs += ["Key", "Entity", "NavigationField"]

    for tree in metadata_trees.values():
        root = tree.getroot()
        for schema in root.findall('.//{http://schemas.microsoft.com/ado/2008/09/edm}Schema'):
            if schema.attrib.get('Namespace') == 'SFOData':
                for et in schema.findall('EntityType', ns):
                    entity_name = et.attrib.get('Name', 'Null')
                    key_names = set()
                    key = et.find('Key', ns)
                    if key is not None:
                        key_names = {pr.attrib.get('Name') for pr in key.findall('PropertyRef', ns)}
                    for prop in et.findall('Property', ns):
                        row = {}
                        for attr in all_attrs:
                            if attr in ["Key", "Entity", "NavigationField"]:
                                continue
                            row[attr] = get_attr_value(prop, attr)
                        row["Key"] = "true" if prop.attrib.get("Name") in key_names else "false"
                        row["Entity"] = entity_name
                        row["NavigationField"] = "false"
                        rows.append(row)
                    for nav in et.findall('NavigationProperty', ns):
                        row = {}
                        for attr in all_attrs:
                            if attr in ["Key", "Entity", "NavigationField"]:
                                continue
                            row[attr] = get_attr_value(nav, attr)
                        row["Key"] = "false"
                        row["Entity"] = entity_name
                        row["NavigationField"] = "true"
                        rows.append(row)

    df_dict = pd.DataFrame(rows, columns=all_attrs)

    # Simple EC Data API Dictionary
    simple_cols = [
        "Entity", "Name", "label", "Type", "Key", "required", "picklist", "MaxLength", 
        "NavigationField", "creatable", "updatable", "visible", "filterable", "sortable", "upsertable"
    ]
    # Ensure columns exist
    for col in simple_cols:
        if col not in df_dict.columns:
            df_dict[col] = ""
            
    df_simple = df_dict[simple_cols].copy()

    # Sorting
    df_simple = df_simple.sort_values(
        by=["Entity", "Name", "Key", "required"],
        ascending=[True, True, False, False],
        key=lambda col: col.map(lambda x: 1 if str(x).lower() == "true" else 0) if col.name in ["Key", "required"] else col
    ).reset_index(drop=True)

    # Save to Excel
    with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
        df_simple.to_excel(writer, sheet_name="Simple EC Data API Dictionary", index=False)
    
    print(f"Dictionary saved to {EXCEL_OUTPUT_PATH}")
    return df_simple

# =============================================================================
# API Query and Parsing
# =============================================================================
def parse_api_response(root_entity, json_data):
    """Parse JSON response to list fields with attributes and derived entity."""
    parsed_fields = []
    
    def traverse(data, path, current_entity):
        if isinstance(data, dict):
            for k, v in data.items():
                new_path = f"{path}.{k}"
                if k == "__metadata":
                    continue
                
                # Determine if this key changes the entity
                # User rule: sub string of the first level like %Nav queried in path
                next_entity = current_entity
                if k.endswith("Nav"):
                    next_entity = k[:-3]
                
                if not isinstance(v, (dict, list)):
                    parsed_fields.append({
                        "Entity": current_entity,
                        "Field": k,
                        "Index Location": new_path,
                        "Sample Value": v
                    })
                else:
                    traverse(v, new_path, next_entity)
        elif isinstance(data, list):
            for i, item in enumerate(data):
                new_path = f"{path}[{i}]" # or [] if we want generic
                traverse(item, new_path, current_entity)

    d_content = json_data.get('d', {})
    if 'results' in d_content:
        results = d_content['results']
        if isinstance(results, list) and results:
            # Just take the first item for documentation
            traverse(results[0], "d.results[]", root_entity)
    elif 'result' in d_content:
        result = d_content['result']
        traverse(result, "d.result", root_entity)
    else:
        traverse(d_content, "d", root_entity)
        
    return parsed_fields

def main():
    print("Starting SuccessFactors API Documentation Generator...")
    
    # 0. Authenticate
    global DYNAMIC_ACCESS_TOKEN
    assertion = get_assertion()
    if assertion:
        DYNAMIC_ACCESS_TOKEN = get_access_token(assertion)
        if DYNAMIC_ACCESS_TOKEN:
            print("Successfully obtained dynamic OAuth 2.0 token.")
        else:
            print("Failed to obtain access token. Will try fallback methods.")
    else:
        print("Failed to obtain assertion. Will try fallback methods.")
    
    # 1. Extract Metadata
    df_dictionary = extract_ec_odata_api_dictionary()
    
    # Create lookup dictionary: (Entity, Name) -> Row Series
    dict_lookup = {}
    field_lookup = {} # Fallback
    
    for idx, row in df_dictionary.iterrows():
        dict_lookup[(row['Entity'], row['Name'])] = row
        field_lookup[row['Name']] = row 
        
    print("Step 1: Metadata Extraction Completed.")

    # 2. Query API and Generate Sheets
    if not os.path.exists(INTEGRATION_FILE):
        print(f"Template file {INTEGRATION_FILE} not found.")
        return

    # Create output file from template
    output_filename = "SF_API_Documentation_Generated.xlsx"
    
    # Load template to copy from
    template_wb = openpyxl.load_workbook(INTEGRATION_FILE)
    template_ws = template_wb["API Template"]
    
    # Load Master Table List for Endpoints and Metadata
    queries_to_run = []
    try:
        master_ws_template = template_wb["SF Master Table List"]
        # Iterate rows starting from row 2 (skip header)
        for i, row in enumerate(master_ws_template.iter_rows(min_row=2, values_only=True), start=2):
            # Col A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11, M=12, N=13, O=14
            # We need:
            # API Name: Col B (idx 1)
            # Entity: Col C (idx 2)
            # Introduction: Col E (idx 4)
            # Endpoint: Col F (idx 5)
            # Trigger Point: Col I (idx 8)
            # Data Flow: Col J (idx 9)
            # System: Col N (idx 13)
            # Category: Col O (idx 14)
            
            if len(row) > 14:
                system_val = row[13]
                category_val = row[14]
                
                # Filter: Only query where System is SuccessFactors and Category is API Resource
                if system_val == "SuccessFactors" and category_val == "API Resource":
                    ent = row[2]
                    endpoint = row[5]
                    
                    if ent and endpoint:
                        queries_to_run.append({
                            "row_idx": i,
                            "entity": ent,
                            "api_name": row[1],
                            "intro": row[4],
                            "endpoint": endpoint,
                            "trigger": row[8],
                            "data_flow": row[9]
                        })
    except Exception as e:
        print(f"Warning: Could not load SF Master Table List: {e}")

    # Create new workbook for output
    template_wb.save(output_filename)
    wb = openpyxl.load_workbook(output_filename)
    template_ws = wb["API Template"]
    output_master_ws = wb["SF Master Table List"]

    used_dict_keys = set()

    for metadata in queries_to_run:
        entity = metadata["entity"]
        print(f"Processing {entity} ({metadata['api_name']})...")
        
        template_url = metadata["endpoint"]
        
        # Use template from Master Table
        # It might contain {Test_API-Server} placeholder
        today_str = datetime.now().strftime('%Y-%m-%d')
        endpoint = template_url.replace("{{Test_API-Server}}", API_SERVER)\
                               .replace("{Test_API-Server}", API_SERVER)\
                               .replace("{today}", today_str)
            
        # Query API
        try:
            resp = make_request(endpoint)
            resp.raise_for_status()
            json_data = resp.json()
        except Exception as e:
            print(f"Error querying {entity}: {e}")
            json_data = {}
            
        # Parse Response
        parsed_fields = parse_api_response(entity, json_data)
        
        # Create Sheet
        sheet_name_raw = metadata["api_name"] if metadata["api_name"] else entity
        sheet_name = sheet_name_raw[:31] # Max 31 chars
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        ws = wb.copy_worksheet(template_ws)
        ws.title = sheet_name
        print(f"Created sheet: {sheet_name}")
        
        # Limit JSON data for documentation (Top 3 items in lists)
        def limit_json_data(data, limit=3):
            if isinstance(data, dict):
                return {k: limit_json_data(v, limit) for k, v in data.items()}
            elif isinstance(data, list):
                return [limit_json_data(i, limit) for i in data[:limit]]
            else:
                return data
        
        json_data_limited = limit_json_data(json_data)
        json_str = json.dumps(json_data_limited, indent=4)
        
        # Fill Header Info in Entity Sheet
        ws["A1"] = metadata["api_name"]   # API Name
        ws["B2"] = metadata["data_flow"]  # Data Flow
        ws["B3"] = metadata["trigger"]    # Trigger Point
        ws["B4"] = entity                 # Start Entity
        ws["B5"] = metadata["intro"]      # Introduction
        ws["B6"] = endpoint               # Test API Endpoint
        
        # Fill Sample Response (B12)
        ws["B12"] = json_str
        
        # Write Sample Response to SF Master Table List (Column L -> 12)
        output_master_ws.cell(row=metadata["row_idx"], column=12, value=json_str)
        
        # Fill Fields
        start_row = 14
        for i, field_data in enumerate(parsed_fields):
            row_num = start_row + i
            field_name = field_data["Field"]
            index_loc = field_data["Index Location"]
            field_entity = field_data["Entity"]
            sample_val = field_data["Sample Value"]
            
            # Lookup attributes using derived entity
            meta = dict_lookup.get((field_entity, field_name))
            if meta is None:
                # Fallback to field lookup if exact match fails
                meta = field_lookup.get(field_name)
            
            label = meta["label"] if meta is not None else ""
            type_ = meta["Type"] if meta is not None else ""
            max_len = meta["MaxLength"] if meta is not None else ""
            picklist = meta["picklist"] if meta is not None else ""
            key = meta["Key"] if meta is not None else ""
            required = meta["required"] if meta is not None else ""

            # Standardize logic: Add to used keys to prevent deletion in cleanup step
            # We add both the derived entity/field and the matched metadata entity/field (if any)
            # to ensure we preserve the relevant rows in the master dictionary even if naming varies slightly.
            used_dict_keys.add((field_entity, field_name))
            if meta is not None:
                used_dict_keys.add((meta["Entity"], meta["Name"]))
            
            # Update Column Mapping based on Row 13 of Template
            # Field: B (2)
            # Entity: C (3)
            # Path: D (4)
            # Sample Value: E (5)
            # Label: F (6)
            # Type: G (7)
            # Key: H (8)
            # Required: I (9)
            # Picklist: J (10)
            # MaxLength: K (11)
            
            ws.cell(row=row_num, column=2, value=field_name)
            ws.cell(row=row_num, column=3, value=field_entity) # Entity in C
            ws.cell(row=row_num, column=4, value=index_loc)    # Path in D
            ws.cell(row=row_num, column=5, value=str(sample_val) if sample_val is not None else "") # Sample Value in E
            ws.cell(row=row_num, column=6, value=label)
            ws.cell(row=row_num, column=7, value=type_)
            ws.cell(row=row_num, column=8, value=key)
            ws.cell(row=row_num, column=9, value=required)
            ws.cell(row=row_num, column=10, value=picklist)
            ws.cell(row=row_num, column=11, value=max_len)
            
            ws.cell(row=row_num, column=11, value=max_len)
    
    # =============================================================================
    # Cleanup Unused Fields from Dictionaries
    # =============================================================================
    print("Cleaning up unused fields in Master Dictionary and Dropdown Mapping...")

    # 1. Cleanup SF Master Data Dictionary
    if "SF Master Data Dictionary" in wb.sheetnames:
        ws_dict = wb["SF Master Data Dictionary"]
        rows = list(ws_dict.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            data = rows[1:]
            try:
                # Identify columns. Assuming Entity is A and Field is B based on user info.
                # However, robustly finding them by name is better.
                # Standard headers: Entity, Field
                ent_idx = header.index("Entity")
                field_idx = header.index("Field")
                
                filtered_rows = [header]
                kept_count = 0
                for r in data:
                    if len(r) > max(ent_idx, field_idx):
                        key = (r[ent_idx], r[field_idx])
                        if key in used_dict_keys:
                            filtered_rows.append(r)
                            kept_count += 1
                
                # Clear and Write back
                ws_dict.delete_rows(1, amount=ws_dict.max_row)
                for r in filtered_rows:
                    ws_dict.append(r)
                print(f"SF Master Data Dictionary cleaned. Kept {kept_count} rows.")
            except ValueError:
                print("Warning: 'Entity' or 'Field' column not found in SF Master Data Dictionary.")

    # 2. Cleanup SF DropdownList Mapping
    if "SF DropdownList Mapping" in wb.sheetnames:
        ws_drop = wb["SF DropdownList Mapping"]
        rows = list(ws_drop.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            data = rows[1:]
            try:
                # Standard headers: Entity, Name
                ent_idx = header.index("Entity")
                name_idx = header.index("Name")
                
                filtered_rows = [header]
                kept_count = 0
                for r in data:
                    if len(r) > max(ent_idx, name_idx):
                        key = (r[ent_idx], r[name_idx])
                        if key in used_dict_keys:
                            filtered_rows.append(r)
                            kept_count += 1
                
                # Clear and Write back
                ws_drop.delete_rows(1, amount=ws_drop.max_row)
                for r in filtered_rows:
                    ws_drop.append(r)
                print(f"SF DropdownList Mapping cleaned. Kept {kept_count} rows.")
            except ValueError:
                print("Warning: 'Entity' or 'Name' column not found in SF DropdownList Mapping.")

    wb.save(output_filename)
    print(f"Documentation generated: {output_filename}")

if __name__ == "__main__":
    main()
